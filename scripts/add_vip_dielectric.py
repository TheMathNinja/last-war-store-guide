import sys
from pathlib import Path

from openpyxl import load_workbook


def add_vip_dielectric(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active

    vip_start_col = 7
    item_col = vip_start_col
    qty_col = vip_start_col + 1
    price_col = vip_start_col + 2
    curr_col = vip_start_col + 3
    limit_col = vip_start_col + 4

    for row in range(3, ws.max_row + 20):
        item = ws.cell(row=row, column=item_col).value
        if isinstance(item, str) and item.strip().lower() == "dielectric ceramic":
            ws.cell(row=row, column=qty_col).value = 1
            ws.cell(row=row, column=price_col).value = 64
            ws.cell(row=row, column=curr_col).value = "DIA"
            ws.cell(row=row, column=limit_col).value = 50
            wb.save(path)
            return
        if item in (None, ""):
            ws.cell(row=row, column=item_col).value = "Dielectric Ceramic"
            ws.cell(row=row, column=qty_col).value = 1
            ws.cell(row=row, column=price_col).value = 64
            ws.cell(row=row, column=curr_col).value = "DIA"
            ws.cell(row=row, column=limit_col).value = 50
            wb.save(path)
            return

    raise RuntimeError("No open VIP Storefront row found.")


if __name__ == "__main__":
    targets = sys.argv[1:] or ["data/Last War Price Guide.xlsx"]
    for target in targets:
        add_vip_dielectric(Path(target))
