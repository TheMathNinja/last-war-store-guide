from __future__ import annotations

import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


MANAGED_STORES = {
    "Total Mobilization Store": [
        ("UR Hero Universal Shard", 10, 10, "MOB", 1),
        ("Hero Choice Chest", 1, 2, "MOB", 800),
        ("Survivor Recruitment Ticket", 10, 8, "MOB", 50),
        ("Lv 5 Drone Component Chest", 1, 30, "MOB", 50),
        ("UR Gear Blueprint", 1, 12, "MOB", 12),
        ("Premium Chip Material", 100, 7, "MOB", 80),
        ("Hero EXP Chest (SSR)", 1, 1, "MOB", 300),
        ("Upgrade Ore", 120, 1, "MOB", 1000),
        ("Skill Medal", 200, 1, "MOB", 1000),
        ("Resource Choice Chest (SSR)", 1, 1, "MOB", 4000),
    ],
    "ID Points Mall": [
        ("UR Hero Universal Shard", 1, 40, "ID", 10),
        ("Hero EXP Chest (SSR)", 1, 20, "ID", 10),
        ("Emote (Gold Bricks)", 1, 100, "ID", 1),
        ("Iron Chest (SSR)", 1, 20, "ID", 10),
        ("Food Chest (SSR)", 1, 20, "ID", 10),
        ("Coin Chest (SSR)", 1, 20, "ID", 10),
        ("Drone Parts", 1, 20, "ID", 10),
        ("1h Construction Speed-Up", 1, 10, "ID", 10),
        ("1h Research Speed-Up", 1, 10, "ID", 10),
        ("1h Training Speed-Up", 1, 10, "ID", 10),
        ("1h Healing Speed-Up", 1, 10, "ID", 10),
    ],
    "Wandering Merchant": [
        ("UR Hero Universal Shard", 10, 2000, "DIA", None),
        ("Resource Choice Chest (UR)", 10, 1500, "DIA", None),
    ],
}

ADDITIONAL_ROWS = {
    "Bounty Hunter Trade Store": [
        ("Drone Parts", 10, 32, "BOUN", 500),
        ("Valor Badge", 100, 50, "BOUN", 100),
    ],
}


def clean_item(item: object, store: str) -> object:
    if item is None:
        return item

    text = str(item).strip()
    if not text:
        return item

    text = re.sub(r"\bSkil Chip\b", "Skill Chip", text, flags=re.I)
    text = re.sub(r"Dialec?tric|Dialetric", "Dielectric", text, flags=re.I)
    text = text.replace("Â®", "")
    text = text.replace("®", "")
    text = re.sub(r"\s+", " ", text).strip()

    if store == "Campaign Storefront" and re.search(r"Hero EXP Chest", text, flags=re.I):
        text = "Hero EXP Chest (SSR)"

    text = re.sub(r"^SSR Hero EXP Chest$", "Hero EXP Chest (SSR)", text, flags=re.I)
    text = re.sub(r"^SR Hero EXP Chest$", "Hero EXP Chest (SR)", text, flags=re.I)
    text = re.sub(r"^UR Hero EXP Chest$", "Hero EXP Chest (UR)", text, flags=re.I)

    text = re.sub(r"^Universal UR Hero Shard$", "UR Hero Universal Shard", text, flags=re.I)
    text = re.sub(r"^UR Hero Universal Shard$", "UR Hero Universal Shard", text, flags=re.I)

    text = re.sub(r"^SSR Resource Choice Chest$", "Resource Choice Chest (SSR)", text, flags=re.I)
    text = re.sub(r"^UR Resource Choice Chest$", "Resource Choice Chest (UR)", text, flags=re.I)
    text = re.sub(r"^Research Choice Chest \(UR\)$", "Resource Choice Chest (UR)", text, flags=re.I)
    text = re.sub(r"^Research Choice Chest UR$", "Resource Choice Chest (UR)", text, flags=re.I)
    text = re.sub(r"^Resource Choice Chest SSR$", "Resource Choice Chest (SSR)", text, flags=re.I)
    text = re.sub(r"^Resource Choice Chest UR$", "Resource Choice Chest (UR)", text, flags=re.I)

    text = re.sub(r"^Drone Part$", "Drone Parts", text, flags=re.I)

    text = re.sub(
        r"^Level\s+(\d+)\s+(?:Drone\s+)?Component\s+Choice\s+Chest$",
        r"Lv \1 Drone Component Choice Chest",
        text,
        flags=re.I,
    )
    text = re.sub(
        r"^Level\s+(\d+)\s+(?:Drone\s+)?Component\s+Chest$",
        r"Lv \1 Drone Component Chest",
        text,
        flags=re.I,
    )
    text = re.sub(
        r"^Lv\.?\s*(\d+)\s+(?:Drone\s+)?Component\s+Choice\s+Chest$",
        r"Lv \1 Drone Component Choice Chest",
        text,
        flags=re.I,
    )
    text = re.sub(
        r"^Lv\.?\s*(\d+)\s+(?:Drone\s+)?Component\s+Chest$",
        r"Lv \1 Drone Component Chest",
        text,
        flags=re.I,
    )

    text = re.sub(r"^1\s+Hour\s+Universal\s+Speed-?Up$", "1h Universal Speed-Up", text, flags=re.I)
    text = re.sub(r"Speed-up", "Speed-Up", text, flags=re.I)
    text = re.sub(r"1h\s+Healing", "1h Healing", text, flags=re.I)

    return text


def normalize_battle_data_row(item: object, qty: object) -> tuple[object, object]:
    text = "" if item is None else str(item).strip()
    numeric_qty = None
    try:
        numeric_qty = float(qty)
    except (TypeError, ValueError):
        pass

    if re.match(r"^Battle Data$", text, flags=re.I) and numeric_qty == 100000:
        return "Battle Data (100K)", 1
    if re.match(r"^Battle Data$", text, flags=re.I) and numeric_qty == 10000:
        return "Battle Data (10k)", 1
    if re.match(r"^Battle Data \(10K\)$", text, flags=re.I) and numeric_qty == 10:
        return "Battle Data (10k)", 1

    return item, qty


def copy_cell_format(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)


def upsert_store_row(ws, start: int, row_values: tuple[object, object, object, object, object]) -> None:
    item, qty, price, curr, limit = row_values
    target_row = None

    for row in range(3, ws.max_row + 1):
        row_item = ws.cell(row, start).value
        row_qty = ws.cell(row, start + 1).value
        row_price = ws.cell(row, start + 2).value
        row_curr = ws.cell(row, start + 3).value
        if row_item == item and row_qty == qty and row_price == price and row_curr == curr:
            target_row = row
            break

    if target_row is None:
        target_row = 3
        while ws.cell(target_row, start).value not in (None, ""):
            target_row += 1
        template_row = max(3, target_row - 1)
        for offset in range(5):
            copy_cell_format(ws.cell(template_row, start + offset), ws.cell(target_row, start + offset))

    for offset, value in enumerate(row_values):
        ws.cell(target_row, start + offset).value = value


def update_workbook(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active

    starts = [
        col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value and ws.cell(2, col).value == "Item"
    ]

    for start in starts:
        raw_store = str(ws.cell(1, start).value)
        store = raw_store.replace("Invation", "Invasion")
        ws.cell(1, start).value = store

        for row in range(3, ws.max_row + 1):
            item_cell = ws.cell(row, start)
            item_cell.value = clean_item(item_cell.value, store)
            item_cell.value, ws.cell(row, start + 1).value = normalize_battle_data_row(
                item_cell.value,
                ws.cell(row, start + 1).value,
            )

            if store == "VIP Storefront" and item_cell.value == "Hero EXP Chest (UR)":
                ws.cell(row, start + 2).value = 480

            if store == "VIP Storefront" and item_cell.value == "Stamina":
                ws.cell(row, start + 2).value = 60

            if (
                store == "Bounty Hunter Trade Store"
                and item_cell.value == "Lv 5 Drone Component Chest"
                and ws.cell(row, start + 2).value == 220
            ):
                item_cell.value = "Lv 5 Drone Component Choice Chest"

            if store == "Zombie Invasion Store":
                if item_cell.value == "Hero Recruitment Ticket":
                    ws.cell(row, start + 4).value = 10
                if re.search(r"Drone Component Chest$", str(item_cell.value), flags=re.I):
                    ws.cell(row, start + 4).value = 3
                if item_cell.value in {"Coin Chest (SSR)", "SSR Coin Chest"}:
                    ws.cell(row, start + 4).value = 100

        for row_values in ADDITIONAL_ROWS.get(store, []):
            upsert_store_row(ws, start, row_values)

    starts = [
        col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value and ws.cell(2, col).value == "Item"
    ]

    headers = ["Item", "Qty", "Price", "Curr", "Limit"]
    template = max(starts)

    for store_name, store_rows in MANAGED_STORES.items():
        starts = [
            col
            for col in range(1, ws.max_column + 1)
            if ws.cell(1, col).value and ws.cell(2, col).value == "Item"
        ]
        existing_store = [col for col in starts if ws.cell(1, col).value == store_name]
        if existing_store:
            start = existing_store[0]
            for row in range(3, ws.max_row + 1):
                for col in range(start, start + 5):
                    ws.cell(row, col).value = None
        else:
            start = max(starts) + 6

        for offset in range(5):
            ws.column_dimensions[ws.cell(1, start + offset).column_letter].width = ws.column_dimensions[
                ws.cell(1, template + offset).column_letter
            ].width
            for row in range(1, ws.max_row + 1):
                copy_cell_format(ws.cell(row, template + offset), ws.cell(row, start + offset))

        ws.cell(1, start).value = store_name
        for offset, header in enumerate(headers):
            ws.cell(2, start + offset).value = header

        for idx, row_values in enumerate(store_rows, start=3):
            for offset, value in enumerate(row_values):
                ws.cell(idx, start + offset).value = value

    wb.save(path)


def main() -> None:
    for arg in sys.argv[1:]:
        update_workbook(Path(arg))
        print(f"updated {arg}")


if __name__ == "__main__":
    main()
